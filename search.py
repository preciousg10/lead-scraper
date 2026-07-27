import requests
import json
import os
import time
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime, timezone, timedelta
import schedule
import shutil
import re
from bs4 import BeautifulSoup

API_KEY = os.environ.get("GOOGLE_PLACES_API_KEY", "")
LAT_START = 41.6
LAT_END = 44.8
LON_START = -83.5
LON_END = -74.5
GRID_STEP_LAT = 0.072
GRID_STEP_LON = 0.096
SEARCH_RADIUS = 8000

NORTHERN_CITIES = [
    "Thunder Bay", "Sudbury", "Sault Ste. Marie", "Timmins",
    "North Bay", "Kenora", "Kirkland Lake", "Kapuskasing",
    "Elliot Lake", "Dryden", "Fort Frances", "Sioux Lookout",
    "Cochrane", "Hearst", "Parry Sound", "Huntsville",
    "Gravenhurst", "Bracebridge", "Haliburton", "Bancroft"
]

SEARCH_TERMS = [
    "plumber",
    "emergency plumber",
    "drain cleaning",
    "plumbing company",
    "water heater repair"
]

MAX_REVIEWS = 125
SEEN_FILE = "seen.json"
OUTPUT_FILE = "leads.xlsx"
SWEEP_INTERVAL_HOURS = 2
TWO_YEARS_AGO = datetime.now(timezone.utc) - timedelta(days=730)

OWNER_KEYWORDS = [
    "owner", "founder", "president", "ceo", "principal",
    "proprietor", "established by", "started by", "founded by",
    "operated by", "run by", "managed by"
]

OWNER_PAGES = ["about", "about-us", "our-story", "team", "contact", "who-we-are"]

def generate_grid():
    points = []
    lat = LAT_START
    while lat <= LAT_END:
        lon = LON_START
        while lon <= LON_END:
            points.append((round(lat, 4), round(lon, 4)))
            lon += GRID_STEP_LON
        lat += GRID_STEP_LAT
    return points

def get_area_name(lat, lon):
    if lat > 44.2:
        return "Barrie and North"
    elif lat > 43.9:
        if lon > -79.8 and lon < -79.0:
            return "York Region"
        elif lon > -80.6 and lon < -79.8:
            return "Guelph-Waterloo Region"
        elif lon < -80.6:
            return "Owen Sound Area"
        else:
            return "Durham and East"
    elif lat > 43.5:
        if lon > -79.7 and lon < -79.1:
            return "Toronto"
        elif lon > -80.3 and lon < -79.7:
            return "Mississauga-Brampton"
        elif lon > -81.0 and lon < -80.3:
            return "Hamilton Area"
        elif lon > -81.5 and lon < -81.0:
            return "Brantford Area"
        elif lon < -81.5:
            return "London Area"
        else:
            return "Oshawa-Whitby Area"
    elif lat > 43.1:
        if lon > -79.5:
            return "Niagara Region"
        elif lon > -81.5:
            return "Cambridge-Kitchener"
        elif lon > -82.5:
            return "Sarnia-Chatham Area"
        else:
            return "Windsor Area"
    else:
        return "Southwest Ontario"

def load_seen():
    if os.path.exists(SEEN_FILE):
        with open(SEEN_FILE, "r") as f:
            return json.load(f)
    return []

def save_seen(seen):
    with open(SEEN_FILE, "w") as f:
        json.dump(seen, f)

def parse_review_time(publish_time_str):
    try:
        return datetime.fromisoformat(publish_time_str.replace("Z", "+00:00"))
    except:
        return None

def get_places(lat, lon, search_term, by_city=False, city_name=""):
    url = "https://places.googleapis.com/v1/places:searchText"
    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": API_KEY,
        "X-Goog-FieldMask": "places.id,places.displayName,places.rating,places.userRatingCount,places.websiteUri,places.nationalPhoneNumber,places.reviews"
    }
    if by_city:
        body = {
            "textQuery": search_term + " in " + city_name + " Ontario Canada",
            "maxResultCount": 20
        }
    else:
        body = {
            "textQuery": search_term + " Ontario Canada",
            "maxResultCount": 20,
            "locationBias": {
                "circle": {
                    "center": {"latitude": lat, "longitude": lon},
                    "radius": SEARCH_RADIUS
                }
            }
        }
    for attempt in range(3):
        try:
            response = requests.post(url, headers=headers, json=body, timeout=15)
            data = response.json()
            return data.get("places", [])
        except requests.exceptions.Timeout:
            if attempt < 2:
                print("    Timeout, retrying (" + str(attempt + 1) + "/3)...")
                time.sleep(3)
            else:
                print("    Failed after 3 attempts, skipping.")
                return []
        except Exception as e:
            print("    ERROR fetching places: " + str(e))
            return []

def find_owner_name(website_url):
    try:
        base_url = website_url.rstrip("/")
        pages_to_check = [base_url] + [base_url + "/" + p for p in OWNER_PAGES]

        for page_url in pages_to_check:
            try:
                response = requests.get(
                    page_url, timeout=6,
                    headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"},
                    allow_redirects=True
                )
                if response.status_code != 200:
                    continue

                soup = BeautifulSoup(response.text, "html.parser")
                text = soup.get_text(" ", strip=True)
                lines = text.split("\n")

                for i, line in enumerate(lines):
                    line_lower = line.lower()
                    for keyword in OWNER_KEYWORDS:
                        if keyword in line_lower:
                            surrounding = " ".join(lines[max(0, i-1):i+3])
                            name_match = re.search(
                                r'\b([A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,2})\b',
                                surrounding
                            )
                            if name_match:
                                name = name_match.group(1)
                                skip_words = ["About", "Contact", "Services", "Plumbing",
                                            "Company", "Ontario", "Canada", "Toronto",
                                            "Call", "Email", "Our", "Your", "The"]
                                if not any(w == name.split()[0] for w in skip_words):
                                    return name

                h_tags = soup.find_all(["h1", "h2", "h3"])
                for h in h_tags:
                    h_text = h.get_text(strip=True)
                    if any(kw in h_text.lower() for kw in OWNER_KEYWORDS):
                        name_match = re.search(
                            r'\b([A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,2})\b',
                            h_text
                        )
                        if name_match:
                            return name_match.group(1)

            except:
                continue

        return "Not found"

    except Exception as e:
        return "Not found"

def check_ads(url):
    try:
        response = requests.get(
            url, timeout=8,
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"},
            allow_redirects=True
        )
        html = response.text.lower()

        google_ads = ["aw-", "google_conversion", "googleadservices.com",
                     "googlesyndication.com", "googleads.g.doubleclick.net"]
        facebook_ads = ["fbq(", "connect.facebook.net/en_us/fbevents", "facebook.com/tr?"]
        other_ads = ["doubleclick.net", "adroll.com"]

        if any(s in html for s in google_ads):
            return True, "Already running Google Ads"
        if any(s in html for s in facebook_ads):
            return True, "Already running Facebook Ads"
        if any(s in html for s in other_ads):
            return True, "Already running retargeting ads"

        has_facebook = "facebook.com" in html
        if has_facebook:
            return False, "No ads - has Facebook page (pitch Meta Ads)"
        return False, "No ads - no Facebook either (pitch Google + Meta)"

    except requests.exceptions.Timeout:
        return False, "Website timed out - likely no ads"
    except Exception:
        return False, "Could not check website"

def load_all_leads():
    if os.path.exists("leads_data.json"):
        with open("leads_data.json", "r") as f:
            return json.load(f)
    return {}

def save_all_leads(all_leads):
    with open("leads_data.json", "w") as f:
        json.dump(all_leads, f)

def rebuild_sheet(all_leads):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Leads"

    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    area_fill = PatternFill(start_color="0f3460", end_color="0f3460", fill_type="solid")
    no_website_fill = PatternFill(start_color="4a0000", end_color="4a0000", fill_type="solid")
    has_website_fill = PatternFill(start_color="0d2b00", end_color="0d2b00", fill_type="solid")

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 25
    ws.column_dimensions["I"].width = 40
    ws.column_dimensions["J"].width = 25
    ws.column_dimensions["K"].width = 30
    ws.column_dimensions["L"].width = 30

    headers = [
        "Name", "Google Maps Link", "Rating", "Reviews",
        "Phone", "Has Website", "Owner Name",
        "On Google Maps Since", "Ads Status",
        "Time in Business", "Last Review Date", "Notes"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for area, leads in all_leads.items():
        if not leads:
            continue

        sorted_leads = sorted(leads, key=lambda x: -x.get("rating", 0))

        ws.append([area] + [""] * 11)
        current_row = ws.max_row
        for col in range(1, 13):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = area_fill
            cell.font = Font(bold=True, color="FFFFFF", size=11)

        for lead in sorted_leads:
            ws.append([
                lead["name"],
                lead["maps_link"],
                lead["rating"],
                lead["review_count"],
                lead["phone"],
                lead["has_website"],
                lead["owner_name"],
                lead["on_maps_since"],
                lead["ads_status"],
                lead["time_in_business"],
                lead["last_review_date"],
                ""
            ])
            current_row = ws.max_row
            ws.cell(row=current_row, column=2).hyperlink = lead["maps_link"]
            ws.cell(row=current_row, column=2).font = Font(color="4fc3f7", underline="single")
            fill = no_website_fill if lead["has_website"] == "No" else has_website_fill
            for col in range(1, 13):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = fill
                if col != 2:
                    cell.font = Font(color="FFFFFF")

    wb.save(OUTPUT_FILE)

def get_archive_name():
    now = datetime.now()
    return "leads-" + now.strftime("%B") + " " + str(now.day) + " " + now.strftime("%Y") + ".xlsx"

def archive_and_reset():
    if os.path.exists(OUTPUT_FILE):
        shutil.copy(OUTPUT_FILE, get_archive_name())
        os.remove(OUTPUT_FILE)
    if os.path.exists("leads_data.json"):
        os.remove("leads_data.json")
    print("Archived. seen.json kept - no business ever re-added.")

def process_place(place, area, all_leads, seen, new_leads):
    place_id = place.get("id", "")
    if place_id in seen:
        return new_leads

    review_count = place.get("userRatingCount", 0)
    reviews = place.get("reviews", [])
    seen.append(place_id)

    if review_count > MAX_REVIEWS:
        return new_leads

    most_recent = None
    oldest = None
    for review in reviews:
        dt = parse_review_time(review.get("publishTime", ""))
        if dt:
            if most_recent is None or dt > most_recent:
                most_recent = dt
            if oldest is None or dt < oldest:
                oldest = dt

    if most_recent and most_recent < TWO_YEARS_AGO:
        name = place.get("displayName", {}).get("text", "N/A")
        print("    Skipping " + name + " - last review too old")
        return new_leads

    last_review_date = most_recent.strftime("%B %d %Y") if most_recent else "No reviews yet"

    if oldest:
        days_old = (datetime.now(timezone.utc) - oldest).days
        years = days_old // 365
        months = days_old // 30
        time_in_business = (str(years) + "+ years") if years > 0 else (str(months) + "+ months")
        on_maps_since = "~" + oldest.strftime("%B %Y")
    else:
        time_in_business = "Unknown"
        on_maps_since = "Unknown"

    name = place.get("displayName", {}).get("text", "N/A")
    website = place.get("websiteUri", "")
    phone = place.get("nationalPhoneNumber", "N/A")
    maps_link = "https://www.google.com/maps/place/?q=place_id:" + place_id
    rating = place.get("rating", 0)
    has_website = "Yes" if website else "No"
    date_found = datetime.now().strftime("%Y-%m-%d")

    owner_name = "No website"
    ads_status = "No website - pitch website + ads package"

    if website:
        print("    Checking ads for: " + name)
        running_ads, ads_status = check_ads(website)
        if running_ads:
            print("    - Skipped " + name + " (" + ads_status + ")")
            return new_leads
        print("    Looking for owner name: " + name)
        owner_name = find_owner_name(website)

    if area not in all_leads:
        all_leads[area] = []

    all_leads[area].append({
        "name": name,
        "maps_link": maps_link,
        "rating": rating,
        "review_count": review_count,
        "phone": phone,
        "has_website": has_website,
        "owner_name": owner_name,
        "on_maps_since": on_maps_since,
        "ads_status": ads_status,
        "time_in_business": time_in_business,
        "last_review_date": last_review_date,
        "date_found": date_found
    })

    new_leads += 1
    print("    + " + name + " | " + str(review_count) + " reviews | " + has_website + " website | " + owner_name + " | " + ads_status)
    return new_leads

def run_sweep():
    print("Starting sweep at " + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    seen = load_seen()
    all_leads = load_all_leads()
    new_leads = 0
    grid = generate_grid()
    print("Grid points: " + str(len(grid)))
    print("Northern cities: " + str(len(NORTHERN_CITIES)))

    for lat, lon in grid:
        area = get_area_name(lat, lon)
        for search_term in SEARCH_TERMS:
            try:
                places = get_places(lat, lon, search_term)
                for place in places:
                    new_leads = process_place(place, area, all_leads, seen, new_leads)
                time.sleep(0.5)
            except Exception as e:
                print("    ERROR: " + str(e))
        save_seen(seen)
        save_all_leads(all_leads)

    for city in NORTHERN_CITIES:
        area = city + " Area"
        for search_term in SEARCH_TERMS:
            try:
                places = get_places(0, 0, search_term, by_city=True, city_name=city)
                for place in places:
                    new_leads = process_place(place, area, all_leads, seen, new_leads)
                time.sleep(0.5)
            except Exception as e:
                print("    ERROR: " + str(e))
        save_seen(seen)
        save_all_leads(all_leads)

    rebuild_sheet(all_leads)
    print("Sweep done. " + str(new_leads) + " new leads added.")
    print("Next sweep in " + str(SWEEP_INTERVAL_HOURS) + " hours...")

def daily_reset():
    print("New day - archiving...")
    archive_and_reset()
    run_sweep()

print("Script started")
run_sweep()
schedule.every(SWEEP_INTERVAL_HOURS).hours.do(run_sweep)
schedule.every().day.at("00:00").do(daily_reset)

while True:
    schedule.run_pending()
    time.sleep(60)